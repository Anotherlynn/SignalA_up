#!./.conda/envs/py310/bin/python3.10

# This is the data crawling file for SignalA_up project
# author: Xinyi Li
# contact: xl4412@nyu.edu
# time of completion: 2023-08-20

import pandas as pd
import numpy
from selenium import webdriver
import time
from fake_useragent import UserAgent
import requests
from bs4 import BeautifulSoup
import openpyxl
import pprint
import datetime
import random
import exchange_calendars as trade_date
from lxml import etree
import xlwt
import xlrd
from xlutils import copy

import jieba
import re

import browser_cookie3
from faker import Factory
import wordcloud
import matplotlib.pyplot as plt
from datetime import date, datetime, timedelta
# requests.packages.urllib3.disable_warnings()
calendar = trade_date.get_calendar("XSHG")

jieba.load_userdict("data/glossary.txt")

def is_chi(wd):
    pattern = re.compile(r'[^\u4e00-\u9fa5]')
    if pattern.search(wd):
        return False
    else:
        return True
#
#
append_list = ['皇庭','国际','沪','深','年','月','日','交易','交易所','一百','公开','信息','显示','证券','券商','搞','华西','公布','上榜','当日','苏常柴']

with open('data/stopwords.txt', 'r') as f:
    stopwords = [line.strip() for line in f]

for i in append_list:
    stopwords.append(i)
def get_word(df):
    # df['title'] = df['title']+df['content']
    word_list =[]
    for i in df['title']:
        for wd in jieba.lcut(i):
            if wd not in stopwords and is_chi(wd):
                word_list.append(wd)
    for i in df['title']:
        for wd in jieba.lcut(i):
            if wd not in stopwords and is_chi(wd):
                word_list.append(wd)
    return word_list

def getHtml():
    res = requests.get("http://127.0.0.1:5555/random").text.strip()
    return res


import requests


# 随机获取一个ip
def get_proxy():
    return requests.get("http://127.0.0.1:5010/get/").json()


# # 删除一个ip
# def delete_proxy(proxy):
#     requests.get("http://127.0.0.1:5010/delete/?proxy={}".format(proxy))
#
#
# def getHtml():
#     retry_count = 5
#     # 获得一个随机ip
#     proxy = get_proxy().get("proxy")
#     while retry_count > 0:
#         try:
#             # 尝试用获得的ip去访问测试网址
#             html = requests.get('http://www.example.com', proxies={"http": "http://{}".format(proxy)})
#             return html
#         # 如果报错，则尝试次数-1，五次后停止访问
#         except Exception:
#             retry_count -= 1
#     # 用完一次就删除代理池中这条代理ip
#     delete_proxy(proxy)
#     return None


# getHtml()
import requests


def get_url(code,page1,page2=200):
    '''
    获取东方财富网股吧链接列表
    code是指公司代码
    page是值爬取页数
    '''
    url_list = []
    for page in range(page1,page2+1):
        url = f"http://guba.eastmoney.com/list,{code},f_{page}.html"
        url_list.append(url)

    return url_list

index=2
def func():
    global index
    index+=1

def get_use_url(url_list,date_1,date_2):
    url_use = []
    proxy = get_proxy().get("proxy")
    proxies = {
            'http': 'http://{}'.format(proxy)
    }
    for i in url_list:
        try:
            # 获得一个随机ip

            headers = {
                'User-Agent': UserAgent().random,
                # 'cookies':  browser_cookie3.edge(domain_name=url)
            }
            res = requests.get(i, headers=headers,proxies=proxies)
            if res.status_code!=200:
                raise Exception('error')
            res.encoding = res.apparent_encoding
            # html_text = bytes(bytearray(res.text, encoding='utf-8'))
            html = etree.HTML(res.text)
            name = html.xpath("/html/head/title/text()")[0][:4]
            # page = html.xpath('//*[@id="mainlist"]/div/ul/li[1]/ul/li[8]/a/span')
            pub1 = html.xpath("//tr[contains(@class,'listitem')][1]//td//div[contains(@class,'update')]//text()")[0]
            pub2 = html.xpath("//tr[contains(@class,'listitem')][last()]//td//div[contains(@class,'update')]//text()")[0]
            pub1 = str('2023-' + str(pub1) + ':00')
            pub2 = str('2023-' + str(pub2) + ':00')
            if name==[] or date_1 > pub1.split()[0]:
                print('error!')
                break
            elif date_2 < pub2.split()[0]:
                continue
            elif date_2 >= pub1.split()[0] and pub2.split()[0]>=date_1:
                print(pub1)
                url_use.append(i)
                print('useful url!')
            # print(pub1)
        except:
            # print(html.xpath("//tr[contains(@class,'listitem')][1]//td//div[contains(@class,'update')]//text()"))
            print('except error!')
            continue
        time.sleep(5)

    return url_use
def get_news(url_list,code):
    '''
    获取新闻列表
    :param url_list:链接列表
    :return: 本地xls
    '''
    #
    # outwb = openpyxl.Workbook()
    # outws = outwb.create_sheet(index=0)
    # outws.cell(row=1,column=1,value='read')
    # outws.cell(row = 1, column = 2, value = "comment")
    # outws.cell(row = 1, column = 3, value = "title")
    # outws.cell(row = 1, column = 4, value = "publ")
    # outws.cell(row = 1, column = 5, value = "url")
    read_llist = []
    comment_llist =[]
    title_llist = []
    publ_llist = []
    url_lllist = []
    # proxies = {
    #     'http': 'http://{}'.format(proxy),
    #     'https': 'https://{}'.format(proxy)}
    for i in range(0,len(url_list)):
        url = url_list[i]
        # 获得一个随机ip
        headers = {
            'User-Agent': UserAgent().random,
            # 'cookies':  browser_cookie3.edge(domain_name=url)
        }
        proxy = get_proxy().get("proxy")
        proxies = {
            'http': 'http://{}'.format(proxy)
        }
        res = requests.get(url, headers=headers,proxies=proxies)
        res.encoding = res.apparent_encoding
        html = etree.HTML(res.text)
        title_list = html.xpath("//tr[contains(@class,'listitem')]//td//div[contains(@class,'title')]//text()")
        read_list = html.xpath("//tr[contains(@class,'listitem')]//td//div[contains(@class,'read')]//text()")
        comment_list = html.xpath("//tr[contains(@class,'listitem')]//td//div[contains(@class,'reply')]//text()")
        # author_list = soup.select("div.author")
        renew_list = html.xpath("//tr[contains(@class,'listitem')]//td//div[contains(@class,'update')]//text()")
        urll_list = html.xpath("//tr[contains(@class,'listitem')]//td//div[contains(@class,'title')]//a/@href")
                # 如果报错，则尝试次数-1，五次后停止访问
        if res.status_code!=200:
            raise Exception('error')

        for k in range(0,len(title_list)):
            print(str('2023-'+str(renew_list[k])+':00'))
            read_llist.append(str(read_list[k]))
            comment_llist.append(str(comment_list[k]))
            title_llist.append(str(title_list[k]))
            publ_llist.append(str('2023-'+str(renew_list[k])+':00'))
            if urll_list[k].startswith('/news,'):
                url_lllist.append('http://guba.eastmoney.com'+str(urll_list[k]))
            else:
                url_lllist.append(str(urll_list[k]))
                # outws.cell(row=index, column=1, value=str(read_list[k]))
                # outws.cell(row=index, column=2, value=str(comment_list[k]))
                # outws.cell(row=index, column=3, value=str(title_list[k]))
                # # outws.cell(row=index, column=4, value=str(author_list[k].text.strip()))
                # outws.cell(row=index, column=4, value=pub_time)
                # outws.cell(row=index,column=5,value='http://guba.eastmoney.com'+str(urll_list[k]))

                # proxy2 = getHtml()
                # proxies2 = {
                #     'http': 'http://{}'.format(proxy2),
                #     'https': 'https://{}'.format(proxy2),
                # }
                # res2 = requests.get(url2,headers=headers,proxies=proxies)
                # time.sleep(3)
                # res2.encoding = res2.apparent_encoding
                # html2 = res2.text
                # soup2 = BeautifulSoup(html2, "html.parser")
                # try:
                #     outws.cell(row=index, column=5, value=str(soup2.select("div.newstext")[0].text.replace('\xa0','').replace('$','')))
                # except:
                #     print('error')
                #     continue
                # func()
            print("第%d条成功写入"%k)

        time.sleep(random.uniform(50,120))

    # outwb.save("data/东方财富网资讯%s.xlsx"%code)
    print("%s saved successfully"%code)

    df = pd.DataFrame(data={'publ':publ_llist,
                           'read':read_llist,
                           'comment':comment_llist,
                           'title':title_llist,
                           'url':url_lllist})
    df['Date'] = df['publ'].apply(lambda x: x.split()[0])
    return df

def get_content(df):
    file = df
    # proxies2 = {
    #     'http': 'http://{}'.format(proxy),
    #     'https': 'https://{}'.format(proxy)}
    for i in file.index:
        headers = {
            'User-Agent': UserAgent().random,
            # 'cookies':  browser_cookie3.edge(domain_name=url)
        }
        res = requests.get(file.loc[i,'url'], headers=headers)
        # 如果报错，则尝试次数-1，五次后停止访问
        if res.status_code != 200:
            file.loc[i,'content'] = 'error'
            print('error')
        else:
            res.encoding = res.apparent_encoding
            html = etree.HTML(res.text)
            file.loc[i,'content'] = "".join(html.xpath("//div[contains(@class,'newstext ')]//p//text()"))
            print('added')
    result_ = df.groupby('Date').apply(get_word).reset_index(drop=True)
    return df,result_


if __name__ == "__main__":
    # tick = pd.read_csv('data/flat_o2v.csv', index_col=0)
    # tick =tick[tick['is_ST']==False]
    # tick = tick[tick.flat_date>20230000]
    # data = tick.groupby('symbol',as_index=False)['flat_date'].min()
    # data['symbol'] = data['symbol'].apply(lambda x: str(x).zfill(6))
    # file = pd.read_csv('data/东方财富网资讯%s.csv'%code, index_col=0)
    # print(file.head())

    tick = pd.read_csv('data/flat_o2v.csv', index_col=0)
    tick = tick[tick['is_ST'] == False]
    tick = tick[tick.flat_date > 20230000]
    tick['10prev'] = tick['flat_date'].apply(lambda x: datetime.strptime(str(x), '%Y%m%d'))
    delta = timedelta(days=11)
    tick['10prev'] = tick['10prev'] - delta
    tick['10prev'] = tick['10prev'].apply(lambda x: str(x).split()[0])
    tick['symbol'] = tick['symbol'].apply(lambda x:str(x).zfill(6))
    tick['flat_'] = tick['flat_date'].apply(lambda x:str(x)[:4]+"-"+str(x)[4:6]+"-"+str(x)[6:])
    use_tick = tick[['symbol', '10prev','flat_']].reset_index(drop=True)
    # code_list = ['000056','000065','000151','000158','000523','000570','000584']
    # page1_list = [42,154,21,103,32,31,156,155,153]
    # page2_list = [49,158,25,175,37,41,161,161,160]
    df_llist = []
    result_list = []
    url_ = []
    # use_tick = use_tick.reset_index(drop=True)[13:]
    for i in range(0,5):
        code = use_tick.loc[i,'symbol']
        print('start: %s'%code)
        url_list = get_url(code,page1=10,page2=300)
        date_1 = use_tick.loc[i, '10prev']
        date_2 = use_tick.loc[i, 'flat_']
        if date_1 <'2023-00-00':
            url_[i] = []
        else:
            url_use_list = get_use_url(url_list,date_1,date_2)
            url_[i]=url_use_list

        with open("data/url.txt", 'w') as f:
            f.writelines(str(url_))
        f.close()
        print('ok!\n')
        time.sleep(100)

        # print(url_list)
        # date_ = data[data.symbol==code]['flat_date'].values[0]
        # proxy = '10.108.1.166:8088'
        df_= get_news(url_use_list,code)
        # df_,result_  = get_content(df_)
        result_ = df_.groupby('Date').apply(get_word)
        df_llist.append(df_)
        result_list.append(result_)
        print("运行完成")
        time.sleep(240)

    #
    #
    # for i, t in use_tick.head(10).iterrows():
    #     code = t.symbol
    #     date_ = t['10prev']
    #
    for i in range(0,data.shape[0]):
        wc = wordcloud.WordCloud(background_color=None,
                                 font_path="SimHei.ttf",
                                 width=2000,height=1500,mode="RGBA",max_words=50).generate(" ".join(data['text'].values[i]))

        plt.imshow(wc, interpolation="bilinear")
        plt.axis("off")
        plt.title(data.index[i])
        plt.savefig('result/wordcloud_all_%d'%i)
        plt.close()


# import json
# with open('data/liang1024_citys.json', 'r',encoding= 'utf-8') as f:
#     # 解析 JSON 数据
#     data = json.load(f)
# # 取出每一个城市的list dic格式
# city_list = data['provinces']
# # 构建空集
# city_name = []
# # 形成每一个城市的list'
# for i in city_list:
#     city = i['citys']
#     city_names = [d['cityName'] for d in city]
#     for a in city_names:
#         city_name.append(a.replace('市',''))
# # 写入txt
# with open('data/cities.txt', 'w', encoding='utf-8') as f:
#     for city in city_name:
#         f.write(city + '\n')
